#!/usr/bin/env python3
"""Build the compact on-device Medicinpriser dataset (skive 2.5).

Two modes:
  --inspect   Download the source, print sheet names + column headers + a few sample rows.
              RUN THIS FIRST to verify the real field names, then fill in COLUMN_MAP below.
  --build     Transform the source into the compact JSON contract the app decodes, gzipped,
              with one (latest) record per varenummer. Asserts the result is < 3 MB.

Privacy note: this produces the dataset that ships to the phone. All lookups happen on-device;
no varenummer is ever sent to a server. See README.md.
"""
from __future__ import annotations

import argparse
import gzip
import io
import json
import re
import sys
from datetime import date

import requests
from openpyxl import load_workbook

# The Medicinpriser workbook is republished every ~14 days at a NEW, dated URL — the previous file
# 404s once superseded (that broke the daily build 2026-07-29). So we DISCOVER the current file from
# the landing page instead of hard-coding it. The page moved esundhed.dk → sundhedsdatabank.dk
# (2025); it links the latest "medicinpriser-udgivet-<DDMMYYYY>.xlsx" on the gopublic CDN.
LANDING_URL = "https://sundhedsdatabank.dk/medicin/medicinpriser"
XLSX_RE = re.compile(r'https://[^\s"\'<>]*medicinpriser-udgivet-\d{6,8}[^\s"\'<>]*\.xlsx', re.IGNORECASE)
UA = {"User-Agent": "erindra-data build (+https://github.com/jesperww/erindra-data)"}
SOURCE_LABEL = "Lægemiddelstyrelsen / Medicinpriser"
MAX_BYTES = 3 * 1024 * 1024  # spec acceptkriterie 5

# TODO: fill in the real column headers from `--inspect` output. Keys are our compact fields;
# Verified against the real "Medicinpriser data" sheet (--inspect, 2026-06-22). Values are the
# source column headers; price/date columns are ignored. 3 rows per product (AIP/AUP/DDD price
# indicators) share the same metadata → deduped to one record per varenummer.
SHEET_NAME = "Medicinpriser data"
COLUMN_MAP = {
    "vnr": "Varenummer",
    "navn": "Lægemiddel",
    "aktivstof": "Indholdsstof",
    "atc": "ATC",
    "form": "Form",
    "styrke": "Styrke",
    "pakning": "Pakning",
}


def current_source_url() -> str:
    """Discover the latest medicinpriser .xlsx URL from the landing page (the URL is dated + rotates)."""
    print(f"Slår aktuel medicinpris-fil op på {LANDING_URL} …", file=sys.stderr)
    resp = requests.get(LANDING_URL, timeout=120, headers=UA)
    resp.raise_for_status()
    matches = XLSX_RE.findall(resp.text)
    if not matches:
        raise SystemExit(
            f"Fandt intet 'medicinpriser-udgivet-*.xlsx'-link på {LANDING_URL}. "
            "Er siden flyttet/omlagt igen? Tjek den i en browser og opdatér LANDING_URL/XLSX_RE."
        )

    # Pick the newest by the DDMMYYYY date in the filename (there is normally only one link).
    def date_key(url: str) -> str:
        m = re.search(r"udgivet-(\d{2})(\d{2})(\d{4})", url)
        return (m.group(3) + m.group(2) + m.group(1)) if m else ""  # → YYYYMMDD, sortable

    url = max(dict.fromkeys(matches), key=date_key)
    print(f"  → {url}", file=sys.stderr)
    return url


def download() -> bytes:
    url = current_source_url()
    print(f"Henter {url} …", file=sys.stderr)
    resp = requests.get(url, timeout=300, headers=UA)
    resp.raise_for_status()
    print(f"  {len(resp.content):,} bytes", file=sys.stderr)
    return resp.content


def inspect(data: bytes) -> None:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n=== Ark: {sheet} ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                print("Kolonner:", [str(c) for c in row])
            elif i <= 3:
                print(f"  Række {i}:", row)
            else:
                break


def build(data: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Fandt ikke arket '{SHEET_NAME}'. Ark: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    idx = {field: header.index(col) for field, col in COLUMN_MAP.items() if col in header}
    if "vnr" not in idx:
        raise SystemExit(
            f"Fandt ikke varenummer-kolonnen '{COLUMN_MAP['vnr']}'. Kør --inspect og ret COLUMN_MAP.\n"
            f"Kolonner i arket: {header}"
        )

    records: dict[str, dict] = {}
    for row in rows:
        vnr_raw = row[idx["vnr"]]
        if vnr_raw is None:
            continue
        vnr = "".join(ch for ch in str(vnr_raw) if ch.isdigit())
        if len(vnr) != 6:
            continue
        record = {}
        for field, col_idx in idx.items():
            if field == "vnr":
                continue
            value = row[col_idx]
            if value is not None and str(value).strip():
                record[field] = str(value).strip()
        # Latest row wins (the sheet is chronological); keep overwriting.
        records[vnr] = record

    dataset = {"version": date.today().isoformat(), "source": SOURCE_LABEL, "records": records}
    raw = json.dumps(dataset, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    packed = gzip.compress(raw, compresslevel=9)
    print(f"{len(records):,} varenumre · {len(raw):,} bytes JSON · {len(packed):,} bytes gzip",
          file=sys.stderr)
    if len(packed) > MAX_BYTES:
        raise SystemExit(f"Datasæt {len(packed):,} B > loft {MAX_BYTES:,} B — stram feltvalg/dedup.")
    return packed


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--inspect", action="store_true", help="print sheets/columns/sample rows")
    ap.add_argument("--build", action="store_true", help="build the compact gzipped dataset")
    ap.add_argument("--out", default="medicinpriser.json.gz")
    args = ap.parse_args()

    if not (args.inspect or args.build):
        ap.error("angiv --inspect eller --build")

    data = download()
    if args.inspect:
        inspect(data)
    if args.build:
        packed = build(data)
        with open(args.out, "wb") as f:
            f.write(packed)
        print(f"Skrev {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
